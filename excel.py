from openpyxl import *
from openpyxl.styles import PatternFill, Border, Side

import datetime

from competences import *

tab_jour = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi"]

def get_periode(index):
	if index==0: return "Matin"
	elif index==1: return "Après-midi"
	else: return "Nuit" 

def get_color(nom, liste_employes):

	for e in liste_employes:

		if e.nom == nom:

			if not e.regime == "N":
				return "FFFFBF00"
			elif e.est_interimaire:
				return "FFFFFF00"
			else :
				if e.equipe == "ROUGE":
					return "FFFF6D6D"
				elif e.equipe == "VERTE":
					return "FF99FF99"
				elif e.equipe == "BLEUE":
					return "FF00FFFF"



def get_excel_file(planning: list, conges: list, dates: list):
	

	liste_employes = get_competences("Matrice de polyvalence.xlsx")


	jour  = dates[0]
	mois  = dates[1]
	annee = dates[2] 
	semaine = dates[3]	

	side = Side(border_style="thin", color="FF000000")
	border = Border(left=side,
					right=side,
					top=side,
					bottom=side)

	wb = Workbook()
	ws = wb.active


	jour = tab_jour[datetime.datetime(annee, mois, jour).weekday()].upper()


	ws.cell(row=1, column=1, value=f"Semaine {semaine}")
	ws.cell(row=2, column=1, value=jour)

	for (index, _periode) in enumerate(planning): #_periode = (periode, equipe)
		periode = _periode[0]
		equipe  = _periode[1]

		cell = ws.cell(column=3+index, row=2, value=get_periode(index))
		cell.border = border

		current_row = 4

		for machine in get_liste_machines("Matrice de polyvalence.xlsx"):

			nombre_de_postes = int(machine[1])
			nom_machine = machine[4:].lower()

			cell = ws.cell(column=2, row=current_row, value=nom_machine.upper())
			cell.border = border

			if nom_machine not in periode:
				current_row+=4 # 3 postes + 1 de saut
				continue
			else: 
				for i in range(3):

					postes = periode[nom_machine]
					try: 
						nom=postes[i][1]
					except:
						current_row+=1
						continue

					cellule = ws.cell(column=3+index, row=current_row, value=nom)
					current_row+=1


					if nom=="Intérimaire":
						color = "FF00FFFF"
					elif nom=="Poste non affecté":
						color = "FFFF0000"
					else:
						color = get_color(nom, liste_employes)

					fill = PatternFill(fill_type='solid', start_color=color, end_color=color)
					cellule.fill=fill
					cellule.border = border

			current_row+=1


	ws.cell(row=2, column=7, value="Nom")
	ws.cell(row=2, column=8, value="Raison")

	row = 4

	for (index, e) in enumerate(conges):

		cellule_nom    = ws.cell(row=row, column=7, value=e[0])
		cellule_raison = ws.cell(row=row, column=8, value=e[1])

		cellule_raison.border = border
		cellule_nom.border    = border

		cellule_nom.fill = PatternFill(fill_type='solid', start_color=e[2], end_color=e[2])

		row += 2


	row += 4

	ws.cell(row=row, column=7, value="Personnel non affecté")

	liste_employes_disponibles = planning[0][0]["Employés disponibles"]+ \
								 planning[1][0]["Employés disponibles"]+ \
								 planning[2][0]["Employés disponibles"]

	for e in liste_employes_disponibles:
		row+=2
		cellule=ws.cell(row=row, column=7, value=e)
		cellule.border=border
		color=get_color(e, liste_employes)
		cellule.fill=PatternFill(fill_type='solid', start_color=color, end_color=color)


	for col in ws.columns:
		max_length = 0
		column = col[0].column_letter # Get the column name
		for cell in col:
			try: # Necessary to avoid error on empty cells
				if len(str(cell.value)) > max_length:
					max_length = len(str(cell.value))
			except:
				pass
		adjusted_width = (max_length + 2) * 1.2
		ws.column_dimensions[column].width = adjusted_width

	dest_filename = "/home/vincent/Travail/STAGE/programme/planning.xlsx"
	wb.save(filename = dest_filename)
	return dest_filename;