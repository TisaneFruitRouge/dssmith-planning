# -*- coding: utf-8 -*-


from datetime import date
from calendar import monthrange
import openpyxl as pyxl
import json


date_courante = date.today()

annee_courante = date_courante.year
mois_courant = date_courante.month
jour_courant = date_courante.day


with open("conf.json", 'r') as myconf:
	data = myconf.read()

conf = json.loads(data)

FICHIER_DES_CONGES = conf["chemins"]["conges"]

def employe_dans_liste_conge(employe, liste):
	'''
		Cette fonction renvoie Vrai si l'employe est dans la liste fournie en entrée,  faux sinon
	'''
	for nom in liste:
		if employe.nom == nom[0]:
			return True

	return False



def conges(jour: int=jour_courant, mois: int=mois_courant, annee: int=annee_courante, fichier=FICHIER_DES_CONGES):
	'''
		Cette fonction renvoie une liste des employés en congé, avec la raison de leur absence
	'''

	conges = pyxl.load_workbook(fichier)
	
	ws = conges.worksheets[mois-1]
	
	dic_conges = dict()

	nb_jours = monthrange(2021, 2)[1]

	nb_ligne = ws.max_row
	
	liste_conges = []

	for i in range(8, nb_ligne): #on parcours les lignes (employes)
		cellule_nom = ws.cell(column=2, row=i)
		nom = cellule_nom.value
		
		if nom == None: 
			continue
		

		color=cellule_nom.fill.start_color.index

		cellule = ws.cell(column=2+jour, row=i).value

		if (cellule != None):
			liste_conges.append((nom.split("(")[0], cellule))

	return liste_conges


if __name__ == "__main__":
	
	print(conges(1,2,2021))
