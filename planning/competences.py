# -*- coding: utf-8 -*-

from openpyxl import *
from openpyxl.styles import PatternFill
import json

CHEMIN_MATRICE = "/transfert/Matrice de polyvalence.xlsx"

class Employe(object):
	# Classe modélisant les employés
	def __init__(self, nom: str, equipe: str, competences: list, regime: str = "N"):


			self.nom = nom
			self.equipe = equipe

			self.regime = regime

			self.est_disponible = True
			self.poste_occupe = None

			self.est_interimaire = False

			self.competences = competences # [(machine, poste, niveau), ...]

	def __str__(self):
		string = "########################\n"
		string += f"{self.nom} / {self.equipe} / {self.regime}\n"
		string += "Compétences : \n"
		for competence in self.competences:
			string += f" * {competence}\n"
		string+="########################\n"	

		return string

	def nb_competences(self):

		'''
			Cette méthode renvoie le nombre de postes que l'employé peut tenir
		'''

		return len(self.competences)

	def get_competences_machine(self, machine):
		'''
			Cette méthode renvoie la liste des postes que l'employé peut tenir à la machine
			donnée en entrée
		'''
		l = []
		for c in self.competences:
			if c[0] == machine:
				l.append(c)
		return l

	def possede_competence_machine(self, machine):
		'''
			Cette méthode renvoie vrai si l'emplpyé possède une compétence à la machine donnée en
			entrée
		'''
		for c in self.competences:
			if c[0] == machine:
				return True
		return False 

	def tient_poste(self, machine, poste):
		'''
			Cette méthode renvoie vrai si l'employé sait tenir le poste donnée en entée
			à la machine donnée en entrée
		'''
		if self.possede_competence_machine(machine):
			for poste_machine in self.get_competences_machine(machine):
				if poste_machine[1] == poste:
					return True
		return False

	def get_competence_poste(self, machine, poste):
		'''
			Cette méthode renvoie le niveau de compétence de l'employé au poste, fourni en entrée, 
			à la machine fournie en entrée. Elle renvoie 0 si l'employé ne sait pas tenir le poste 
		'''
		if self.possede_competence_machine(machine):
			for poste_machine in self.get_competences_machine(machine):
				if poste_machine[1] == poste:
					return poste_machine[2]
		return 0


def get_liste_machines(chemin_matrice=CHEMIN_MATRICE):
	'''
		Cette fonction renvoie la liste des machine de la matrice de polyvalence
	'''
	matrice_de_polyvalence = load_workbook(chemin_matrice)
	ws = matrice_de_polyvalence.active

	l = []

	for c in range(2, ws.max_column+1):

		nom = ws.cell(row=1, column=c).value

		l.append(nom)

	return list(filter(lambda item: item!=None, l))

def get_nom_machine(poste):
	'''
		Etant donnée un nom de poste (ex: "Conducteur b 1604", "Sous conducteur B2000"),
		cette fonction renvoie le nom de la machine
	'''
	titre_poste = ["palettiseur ", "sous conducteur ", "conducteur ", "prérégleur ", "préparateur"]
	# cas unique ou la "machine" est "préparateur" => on doit gérer ce cas différement des autres
	if poste.lower().replace("préparateur", "").strip() == "":
		return "préparateur"
	elif poste.split(" ")[0].lower() == "contremaître":
		return "contremaîtres"
	else:
		for p in titre_poste:
			poste = poste.replace(p, "", 1)

		return poste.strip().lower()

def get_competences(chemin_matrice=CHEMIN_MATRICE):
	'''
		Cette fonction renvoie la liste des employés (de Class Employé)
	'''
	matrice_de_polyvalence = load_workbook(chemin_matrice) # on ouvre le tableau excel
	liste_employes = list() # liste des employés



	for (index, equipe) in enumerate(matrice_de_polyvalence.worksheets): # on parcours chaque equipe

		ws = matrice_de_polyvalence.worksheets[index]
		
		equipe = ws.title.split(' ')[1] # on recupere le nom de l'equipe (verte, bleue ou rouge)

		for r in range(3, ws.max_row):	# on parcours toutes les lignes
			
			est_interim = False

			employe = ws.cell(column=1, row=r).value
			if (employe==None): continue # si la case est vide


			regime = "N" # on test si l'employé possède un régime spécial

			if (employe[-2] == "P" or employe[-2] == "I"):
				regime = employe[-2]
				employe = employe[:-3]	

			if (employe[-5:-1] == "Int."):
				est_interim = True
				employe = employe[:-6].strip()

			competences = list()
			for c in range(1, ws.max_column): # on parcours toutes les colonnes
				niveau_compétence = ws.cell(column=c, row=r).value
				if (niveau_compétence in [1,2,3,4]): # si la case [c,r] à la valeur 1 alors l'emploé sait conduire ce poste
					
					poste = ws.cell(column=c, row=2).value.lower()
					machine = get_nom_machine(poste)
					poste = poste.replace(machine, "").strip()

					if (machine == "préparateur"): 
						# le "conducteur" de préparateur est appelé "préparateur" et aurait donc eu
						# un poste 2 alors qu'il devrait avoir un poste 0. C'est ce qu'on rectifie ici
						poste = 0
					if (machine == "contremaîtres"):
						if (poste.split(" ")[1].lower()=="intégré"):
							poste = 0
						else:
							poste = 1

					elif poste == "conducteur":
						poste = 0
					elif poste == "sous conducteur":
						poste = 1
					else : 
						poste = 2
					competences.append((machine, poste, niveau_compétence))
				

			e = Employe(employe, equipe, competences, regime) # on créer l'employe et on l'ajoute à la liste
			
			e.est_interimaire = est_interim

			liste_employes.append(e)
	
	return liste_employes


def getkeys(dic):
	'''
		Cette fonction renvoie une liste des clefs du disctionnaire (pas sûr qu'elle soit utile, après coup..)
	'''
	l = []
	
	for key in dic.keys():
		l.append(key)
	return l

if __name__ == '__main__':

	employes = get_competences("Matrice de polyvalence.xlsx")
