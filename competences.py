# -*- coding: utf-8 -*-

from openpyxl import *
from openpyxl.styles import PatternFill
import json


class Employe(object):

	def __init__(self, nom: str, equipe: str, competences: list, regime: str = "N"):


			self.nom = nom
			self.equipe = equipe

			self.regime = regime

			self.est_disponible = True
			self.poste_occupe = None

			self.est_interimaire = False

			self.competences = competences # [(machine, poste, niveau), ...]

	def __str__(self):
		string = "########################\n"
		string += f"{self.nom} / {self.equipe} / {self.regime}\n"
		string += "Compétences : \n"
		for competence in self.competences:
			string += f" * {competence}\n"
		string+="########################\n"	

		return string

	def nb_competences(self):
		return len(self.competences)

	def get_competences_machine(self, machine):
		l = []
		for c in self.competences:
			if c[0] == machine:
				l.append(c)
		return l

	def possede_competence_machine(self, machine):
		for c in self.competences:
			if c[0] == machine:
				return True
		return False 

	def tient_poste(self, machine, poste):
		if self.possede_competence_machine(machine):
			for poste_machine in self.get_competences_machine(machine):
				if poste_machine[1] == poste:
					return True
		return False

	def get_competence_poste(self, machine, poste):
		if self.possede_competence_machine(machine):
			for poste_machine in self.get_competences_machine(machine):
				if poste_machine[1] == poste:
					return poste_machine[2]
		return -1


def get_liste_machines(chemin_tab_excel):
	
	matrice_de_polyvalence = load_workbook(chemin_tab_excel)
	ws = matrice_de_polyvalence.active

	l = []

	for c in range(2, ws.max_column+1):

		nom = ws.cell(row=1, column=c).value

		l.append(nom)

	return list(filter(lambda item: item!=None, l))

def get_nom_machine(poste):
	
	titre_poste = ["palettiseur ", "sous conducteur ", "conducteur ", "prérégleur "]

	for p in titre_poste:
		poste = poste.replace(p, "")

	return poste.strip()

def get_competences(chemin_tab_excel):

	matrice_de_polyvalence = load_workbook(chemin_tab_excel) # on ouvre le tableau excel
	liste_employes = list() # liste des employés



	for (index, equipe) in enumerate(matrice_de_polyvalence.worksheets): # on parcours chaque equipe

		ws = matrice_de_polyvalence.worksheets[index]
		
		equipe = ws.title.split(' ')[1] # on recupere le nom de l'equipe (verte, bleue ou rouge)

		for r in range(3, ws.max_row):	# on parcours toutes les lignes
			
			est_interim = False

			employe = ws.cell(column=1, row=r).value
			if (employe==None): continue # si la case est vide


			regime = "N"

			if (employe[-2] == "P" or employe[-2] == "I"):
				regime = employe[-2]
				employe = employe[:-3]	

			if (employe[-5:-1] == "Int."):
				est_interim = True
				employe = employe[:-6].strip()

			competences = list()
			for c in range(1, ws.max_column): # on parcours toutes les colonnes
				niveau_compétence = ws.cell(column=c, row=r).value
				if (niveau_compétence in [1,2,3,4]): # si la case [c,r] à la valeur 1 alors l'emploé sait conduire ce poste
					
					poste = ws.cell(column=c, row=2).value.lower()
					machine = get_nom_machine(poste)
					poste = poste.replace(machine, "").strip()

					if (machine == "préparateur"): 
						# le "conducteur" de préparateur est appelé "préparateur" et aurait donc eu
						# un poste 2 alors qu'il devrait avoir un poste 0. C'est ce qu'on rectifie ici
						poste = 0

					elif poste == "conducteur":
						poste = 0
					elif poste == "sous conducteur":
						poste = 1
					else : 
						poste = 2
					competences.append((machine, poste, niveau_compétence))
				

			e = Employe(employe, equipe, competences, regime) # on créer l'employe et on l'ajoute à la liste
			
			e.est_interimaire = est_interim

			liste_employes.append(e)
	
	return liste_employes


def getkeys(dic):

	l = []
	
	for key in dic.keys():
		l.append(key)
	return l

if __name__ == '__main__':

	employes = get_competences("Matrice de polyvalence.xlsx")
